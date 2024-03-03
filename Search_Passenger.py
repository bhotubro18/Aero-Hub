import openpyxl as xl

wb = xl.load_workbook("Passenger_Info.xlsx")
ws = wb.active
def search_Passenger():
    columns = []


    max_cols = ws.max_column
    max_rows = ws.max_row
    for cell in ws[1]:
        columns.append(cell.value)


    to_search = input("Enter passport number to search a passenger: ")

    for row in range(1, max_rows + 1):
        for column in range(1, max_cols + 1):
            cell = ws.cell(row, column)
            if cell.value == to_search:
                for i in range(-1, max_cols - 1):
                    print(columns[i+1] + ": " + ws.cell(row = row, column = column + i).value)
        print()       

wb.save("Passenger_Info.xlsx")


