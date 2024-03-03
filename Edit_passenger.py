import openpyxl as xl
import datetime as dt
import Add_Passenger

wb = xl.load_workbook("Passenger_Info.xlsx")
ws = wb.active

list1 = []

max_rows = ws.max_row
max_cols = ws.max_column

def to_edit():
    flag = "1"
    while flag != "0":
        to_edit = input("Enter passport number to be edited: ")
        for row in range(1, max_rows):
            for column in range(1, max_cols):
                 cell = ws.cell(row, 2)
                 if cell.value == to_edit:
                       for i in range(-1, max_cols - 1):
                        ws.delete_rows(row, 1)
                            
                        list1.insert(row, Add_Passenger.add_passenger())
                        break

        flag = input("Press 0 to confirm exit:")                
                
    ws.append(list1)
    wb.save("Passenger_Info.xlsx")



































