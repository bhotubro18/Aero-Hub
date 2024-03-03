import datetime as dt
import openpyxl as xl
def add_passenger():
    
    wb = xl.load_workbook("Passenger_Info.xlsx")
    ws = wb.active

    ws['A1'] = "Date"
    ws['B1'] = "Passport No."
    ws['C1'] = "First Name"
    ws['D1'] = "Middle Name"
    ws['E1'] = "Last Name"
    ws['F1'] = "Flight No Onboard"
    ws['G1'] = "Mobile Number"
    ws['H1'] = "Nationality"

    
    max_rows = ws.max_row
    max_col = ws.max_column

    def add_date():
        date = dt.datetime.now()
        return str(date.strftime("%d-%m-%Y %H:%M:%S"))

    def passport_no():
        pass_no = input("Enter passenger's passport number: ")
        return pass_no

    def add_fname():
        fname = input("Enter passeger's first name: ")
        return fname
    
    def add_mname():
        mname = input("Enter passenger's middle name: ")
        return mname

    def add_lname():
        lname = input("Enter passenger's last name: ")
        return lname

    def add_flight_no():
        flight_no = input("Enter passenger's flight no.: ")
        return flight_no

    def mobile_no():
        mobile_no = input("Enter passenger's mobile number: ")
        return mobile_no

    def nationality():
        nationality = input("Enter passenger's nationality: ")
        return nationality



##############################################################################################################################################


    flag = "1"
    while flag != "0":

        if flag == "0":
            break
        
        else:
            data = []
            details = []

            details.append(add_date())
            details.append(passport_no())
            details.append(add_fname())
            details.append(add_mname())
            details.append(add_lname())
            details.append(add_flight_no())
            details.append(mobile_no())
            details.append(nationality())

            data.append(details)
            for info in data:
                ws.append(details)

            flag = input("To exit press 0 or to continue press any key: ")

    wb.save("Passenger_Info.xlsx")
  
        

