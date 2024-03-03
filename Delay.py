import openpyxl as xl
import datetime as dt

def Delay():
    
    data = []
    details = []

    wb = xl.load_workbook("Delay.xlsx")
    ws = wb.active

    ws['A1'] = "Flight No"
    ws['B1'] = "Hours delayed"
    
    flag = "1"
    
    while (flag != "0"):
        
        if (flag == "0"):
            break
        else:
            flight_no = input("Enter flght no: ")
            hour = int(input("Enter the hours by which the flight is delayed: "))

        details.append(flight_no)
        details.append(hour)

        data.append(details)
        
        for info in data:
            ws.append(details)

        flag = input("Enter 0 to exit or press any key to continue: ")

        wb.save("Delay.xlsx")


















