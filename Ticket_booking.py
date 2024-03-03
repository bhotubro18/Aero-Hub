import openpyxl
from openpyxl import load_workbook
from datetime import date, datetime

def book_ticket():
    # Load the existing workbook
    wb = openpyxl.load_workbook("ticket_booking_data.xlsx")
    sheet = wb.active
    sheet.title = "Sheet1"
    print("active sheet title: " + sheet.title)

    # Set column widths and header values if needed
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    c1 = sheet["A1"]
    c1.value = "From"
    c2 = sheet["B1"]
    c2.value = "To"
    c3 = sheet["C1"]
    c3.value = "Date"
    c4 = sheet["D1"]
    c4.value = "No. of adults (>12)"
    c5 = sheet["E1"]
    c5.value = "No. of children (<12)"
    c6 = sheet["F1"]
    c6.value = "Travel class"
    c7 = sheet["G1"]
    c7.value = "Airlines"
    c8 = sheet["H1"]
    c8.value = "Departure time"

    flag = "1"
    while flag != "0":
        flag = input("To exit press 0 or to continue press any key: ")
        if flag == "0":
            break
        else:
            # Get user input for a new ticket
            fl = []
            tbap = input("From: ")
            tbdp = input("To: ")
            tbyear = int(input('Enter The Year: '))
            tbmonth = int(input('Enter A Month: '))
            tbday = int(input('Enter A Day: '))
            tbda = date(tbyear, tbmonth, tbday)
            tbna = int(input("No. of adults (>12): "))
            tbnc = int(input("No. of children (<12): "))
            tbtc = input("Travel Class: ")
            tbal = input("Airlines name: ")
            tbti = input("Departure time: ")
            fl.extend((tbap, tbdp, tbda, tbna, tbnc, tbtc, tbal, tbti))

            # Append the new ticket to the sheet
            sheet.append(fl)

            # Save the workbook after adding the new ticket
            wb.save("ticket_booking_data.xlsx")

            # Load the newdata workbook and update available tickets
            tbwb = openpyxl.load_workbook("newdata.xlsx")
            sheettb = tbwb.active
            for rows in sheettb.iter_rows(min_row=2, values_only=True):
                if rows[3] == tbap and rows[7] == tbdp and rows[1] == tbal and rows[4] == tbti and rows[8] == tbtc:
                    avai = rows[12]
                    if avai > 0:
                        avai -= 1
                        print("Updated available tickets in ", rows[2], " are: ", avai)
                        sheettb.cell(row=rows[0]+2, column=13, value=avai)  # Assuming column 13 is the column for available tickets
                    elif avai == 0:
                        print("Full")

            # Save the newdata workbook after updating available tickets
            tbwb.save("newdata.xlsx")
            tbwb.close()

    # Close the main workbook after the loop
    wb.close()
