#ticket avail
#tickets in airlines
import openpyxl
from openpyxl import load_workbook

def ticket_available():
    wb=openpyxl.load_workbook("newdata.xlsx")
    ws=wb.active

    ai=10
    aa=10
    gf=10
    ig=10
    sj=10
    vr=10

    #cities and their codes
    ct={0:"Delhi",1:"Mumbai",2:"Kolkata",3:"Hyderabad",4:"Bangalore",5:"Chennai"}
    airl={10:"Air_India",20:"GO_FIRST",30:"Indigo",40:"SpiceJet",50:"Vistara"}
    depti={"a":"Early_Morning","b":"Morning","c":"Afternoon","d":"Evening","e":"Night","f":"Late_Night"}
    cla={"E":"Economy","B":"Business"}

    print("==>For arrival and departure place enter the number shown below  corresponding to cities")
    print("0: Delhi\n1: Mumbai\n2: Kolkata\n3: Hyderabad\n4: Bangalore\n5: Chennai \n")
    arc=int(input("Enter number for departure city: "))
    drc=int(input("Enter number for arrival city : "))

    print("==>For airlines enter the number shown below  corresponding to airlines name")
    print("10: Air_India\n20: GO_FIRST\n30: Indigo\n40: SpiceJet\n50: Vistara\n")
    alp=int(input("Enter number for Airlines name : "))

    print("==>Departure time")
    print("a: Early_Morning\nb: Morning\nc: Afternoon\nd: Evening\ne: Night\nf: Late_Night\n")
    dpt=input("Enter for departure time : ")

    print("==>Class")
    print("E: Economy\nB: Business\n")
    cl=input("Enter for Class : ")


    flag = 1
    for row in ws.iter_rows(min_row=2,values_only=True):
        
        if row[3]==ct[arc] and row[7]==ct[drc] and row[1]==airl[alp] and row[4]==depti[dpt] and row[8]==cla[cl]:
            tolist=list(row)
            print("Total available tickets in flight",row[2],"are : ",tolist[12])
            flag=0

    if flag==1:
        print("No tickets available")


